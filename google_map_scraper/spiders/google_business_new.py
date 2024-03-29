import csv
import logging
import re
from copy import deepcopy
from geopy.geocoders import Nominatim
import scrapy
from urllib.parse import quote_plus, quote
from scrapy import Request
import openpyxl
import usaddress
import random

logger = logging.getLogger(__name__)

shard = 'test'


class GoogleSpider(scrapy.Spider):
    name = "google_business_new"
    custom_settings = {
        'ROBOTSTXT_OBEY': False,
        'ZYTE_SMARTPROXY_ENABLED': True,
        'ZYTE_SMARTPROXY_URL': 'api.zyte.com:8011',
        'ZYTE_SMARTPROXY_APIKEY': 'bf34ea057aa340bfaf06eb0ecb05cc34',
        'CONCURRENT_REQUESTS': 1,
        'CONCURRENT_REQUESTS_PER_DOMAIN': 1,
        'CONCURRENT_ITEMS': 64,
        'AUTOTHROTTLE_ENABLED': True,
        'DOWNLOAD_TIMEOUT': 600,
        'DOWNLOADER_MIDDLEWARES': {
            'scrapy_zyte_smartproxy.ZyteSmartProxyMiddleware': 610
        },
        "X-Crawlera-Region": "CA",
        "HTTPERROR_ALLOW_ALL": True,
        "ITEM_PIPELINES": {
            "google_map_scraper.pipelines.GoogleScraperPipeline": 1,
        },
    }
    us_states = {
        "alabama": "AL",
        "alaska": "AK",
        "arizona": "AZ",
        "arkansas": "AR",
        "california": "CA",
        "colorado": "CO",
        "connecticut": "CT",
        "delaware": "DE",
        "florida": "FL",
        "georgia": "GA",
        "hawaii": "HI",
        "idaho": "ID",
        "illinois": "IL",
        "indiana": "IN",
        "iowa": "IA",
        "kansas": "KS",
        "kentucky": "KY",
        "louisiana": "LA",
        "maine": "ME",
        "maryland": "MD",
        "massachusetts": "MA",
        "michigan": "MI",
        "minnesota": "MN",
        "mississippi": "MS",
        "missouri": "MO",
        "montana": "MT",
        "nebraska": "NE",
        "nevada": "NV",
        "new hampshire": "NH",
        "new jersey": "NJ",
        "new mexico": "NM",
        "new york": "NY",
        "north carolina": "NC",
        "north dakota": "ND",
        "ohio": "OH",
        "oklahoma": "OK",
        "oregon": "OR",
        "tennsylvania": "PA",
        "thode island": "RI",
        "south carolina": "SC",
        "south dakota": "SD",
        "tennessee": "TN",
        "texas": "TX",
        "utah": "UT",
        "vermont": "VT",
        "virginia": "VA",
        "washington": "WA",
        "west virginia": "WV",
        "wisconsin": "WI",
        "wyoming": "WY",
    }
    # new_listings_url_t = "https://www.google.com/localservices/prolist?ssta=1&src=2&scp=CgASABoAKgA%3D&q={q}&lci={page}&ved=0CAUQjdcJahgKEwjw4oSV4bGEAxUAAAAAHQAAAAAQpQE&slp=MgBAAVIECAIgAIgBAA%3D%3D"
    # new_listings_url_t = (
    #     "https://www.google.com/localservices/prolist?ssta=1&src=2&q={q}&lci={page}"
    # )
    new_listings_url_t = (
        "https://www.google.com/localservices/prolist?g2lbs=AP8S6ENgyDKzVDV4oBkqNJyZonhEwT_VJ6_XyhCY8"
        "jgI2NcumLHJ7mfebZa8Yvjyr_RwoUDwlSwZt5ofLQk3D079b7a0tYFMAl-OvnNjzh2HzyjZNDGO0bloXZTJ8ttkCFt5r"
        "wXuqt_u&hl=en-PK&gl=pk&ssta=1&oq={q}&src=2&sa=X&scp=CgASABoAKgA%3D&q={q}&ved=2ahUKEwji7NSKjZ"
        "iAAxUfTEECHdJnDF8QjdcJegQIABAF&slp=MgBAAVIECAIgAIgBAJoBBgoCFxkQAA%3D%3D&lci={page}"
    )
    new_details_url_t = (
        "https://www.google.com/localservices/prolist?g2lbs=AP8S6ENgyDKzVDV4oBkqNJyZonhEwT_VJ6_XyhCY8"
        "jgI2NcumLHJ7mfebZa8Yvjyr_RwoUDwlSwZt5ofLQk3D079b7a0tYFMAl-OvnNjzh2HzyjZNDGO0bloXZTJ8ttkCFt5r"
        "wXuqt_u&hl=en-PK&gl=pk&ssta=1&oq={q}&src=2&sa=X&scp=CgASABoAKgA%3D&q={q}&ved=2ahUKEwji7NSKjZ"
        "iAAxUfTEECHdJnDF8QjdcJegQIABAF&slp=MgBAAVIECAIgAIgBAJoBBgoCFxkQAA%3D%3D&spp={id}"
    )
    RETRY_HTTP_CODES = [400, 403, 407, 408, 429, 500, 502, 503, 504, 405, 503, 520]
    handle_httpstatus_list = [
        400,
        401,
        402,
        403,
        404,
        405,
        406,
        407,
        409,
        500,
        501,
        502,
        503,
        504,
        505,
        506,
        507,
        509,
    ]
    headers = {
        "user-agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.63 Safa"
        "ri/537.36",
        "content-type": "application/json",
        "accept-language": "en-US,en;q=0.9",
        "X-Crawlera-Region": "SE",
    }
    search_keyword = "{keyword} {location} USA"
    scraped_businesses = []
    image_url = "https://www.google.com/_/AdsHomeservicesConsumerUi/data/batchexecute"
    img_data = (
        "f.req=%5B%5B%5B%22wTe8We%22%2C%22%5B%5C%22{}%5C%22%2Cnull%2Cnull%2C%5C%22CgIgAQ%3D%3D%5C%22%2Cnull%2Cn"
        "ull%2Cnull%2Cnull%2Cnull%2Cnull%2Cnull%2Cnull%2C%5B%5C%22{}%5C%22%5D%5D%22%2Cnull%2C%22generic%22%5D%5"
        "D%5D&at=AAuQa1rFVkuavEc7pNrTSCWVJ59L%3A1702712109706&"
    )
    img_header = {
        "accept": "*/*",
        "content-type": "application/x-www-form-urlencoded;charset=UTF-8",
    }

    keywords = []
    locations = []

    def __init__(self):
        super(GoogleSpider, self).__init__()
        with open(f"input/keyword_{shard}.csv", mode="r") as f:
            self.keywords = list(csv.DictReader(f))

        wb_obj = openpyxl.load_workbook("input/test_zipcode.xlsx")
        sheet_obj = wb_obj["USA ALL"]
        columns = ["Zip Code", "City", "County", "Population", "State"]
        # for row in range(2, sheet_obj.max_row + 1):
        #     row_data = {}
        #     for col, col_name in enumerate(columns, start=1):
        #         cell_value = sheet_obj.cell(row=row, column=col).value
        #         row_data[col_name] = cell_value
        #     self.locations.append(row_data)
        self.locations.append({
            'Zip Code': '-',
            'City': 'Maryland',
        })

    def start_requests(self):
        for location in self.locations:
            random.shuffle(self.keywords)
            for keyword in self.keywords:
                query = self.search_keyword.format(
                    keyword=keyword["Keywords"], location=location["City"]
                )
                url = self.new_listings_url_t.format(q=quote_plus(query), page=0)
                meta = {
                    "keyword": keyword["Keywords"],
                    "start": 0,
                    "query": query,
                    "scraped": 0,
                    "duplicate": 0,
                }
                yield Request(
                    url=url,
                    callback=self.parse,
                    headers=self.headers,
                    meta={
                        **meta,
                        "zyte_api_automap": {
                            "geolocation": "SE",
                            "responseCookies": True,
                        },
                    },
                )

    def parse(self, response, *args):
        logger.info(f'delay: {response.meta.get("download_latency", 0)}')
        fb = len(response.css('div[jscontroller="xkZ6Lb"]'))
        logging.info(
            f"found {fb} businesses for {response.meta['keyword']}, {response.meta['start']}"
        )
        for t in response.css(".AIYI7d::text"):
            print(t.get())
        for listing_selector in response.css('div[jscontroller="xkZ6Lb"]'):
            listing_id = (
                listing_selector.css("::attr(data-profile-url-path)")
                .get("")
                .replace("/localservices/profile?spp=", "")
            )
            in_type = listing_selector.css("span.hGz87c::text").get("").strip()
            response.meta["type"] = in_type
            details_url = self.new_details_url_t.format(
                q=quote_plus(response.meta["keyword"]), id=listing_id
            )
            Name = listing_selector.css(".xYjf2e::text").get("").strip()
            Address = (
                listing_selector.css(".hGz87c span::text").getall()[-1]
                if listing_selector.css(".hGz87c span::text").getall()
                else ""
            )
            industry = (
                listing_selector.xpath(".//span[@class='hGz87c']/text()")
                .get("")
                .strip()
            )
            response.meta["Industry"] = industry
            response.meta["feature_id"] = listing_selector.css(
                "::attr(data-feature-id)"
            ).get("")
            if f"{Name} {Address}" not in self.scraped_businesses:
                self.scraped_businesses.append(f"{Name} {Address}")
                response.meta["scraped"] = response.meta.get("scraped", 0) + 1
                logger.info(
                    f'going into details: {response.meta["keyword"]}, scraped: {response.meta["scraped"]}, duplicate: {response.meta["duplicate"]}'
                )
                yield scrapy.Request(
                    url=details_url,
                    callback=self.parse_new_details,
                    meta={**response.meta, 'dont_retry': True},
                    headers=self.headers,
                )
            else:
                response.meta["duplicate"] = response.meta.get("duplicate", 0) + 1
                logger.info(
                    f'Aleady Scraped: {Name} {Address}, {response.meta["keyword"]}, scraped: {response.meta["scraped"]}, duplicate: {response.meta["duplicate"]}'
                )

        if response.css('button[aria-label="Next"]'):
            keyword = response.meta["keyword"]
            start = response.meta["start"] + fb
            query = response.meta["query"]
            url = self.new_listings_url_t.format(q=quote_plus(query), page=start)
            meta = {
                **response.meta,
                "keyword": keyword,
                "start": start,
                "query": query,
                "zyte_api_automap": {
                    "geolocation": "SE",
                },
            }
            logger.info(f"Next Page: {url}")
            yield Request(
                url=url,
                callback=self.parse,
                meta=meta,
                headers=self.headers
                # errback=self.skip_request,
            )

    def skip_request(self, response):
        print(response.meta.items())
        keyword = response.meta["keyword"]
        start = response.meta["start"] + 1
        query = response.meta["query"]
        url = self.new_listings_url_t.format(q=quote_plus(query), page=start)
        meta = {**response.meta, "keyword": keyword, "start": start, "query": query}
        logger.info(f"Skip a page: {url}")
        yield Request(url=url, callback=self.parse, meta=meta, headers=self.headers)
        
    def error(self, failure):
        return failure.request.cb_kwargs.get('item', {})

    def parse_new_details(self, response):
        item = dict()
        item['Keyword'] = response.meta.get('keyword', '')
        item["Bussiness_Name"] = response.css("div.tZPcob::text").get("").strip()
        item["Bussiness_Contact"] = (
            response.css("div.eigqqc::text")
            .get("")
            .strip()
            .replace(" ", "")
            .replace("-", "")
            .replace("+", "")
            .replace(" ", "")
            .replace("-", "")
            .replace("(", "")
            .replace(")", "")
        )
        item["Business_URL"] = response.url
        item["Bussiness_Website"] = (
            response.css("a.iPF7ob::attr(href)")
            .get("")
            .strip()
            .replace("/url?sa=i&source=web&rct=j&", "")
            .replace("url=", "")
        )
        item["Bussiness_Service"] = response.xpath(
            '//*[contains(text(), "Services:")]/following::text()[1]'
        ).get("")
        item["Bussiness_Serving_Area"] = ", ".join(
            response.css("div.oR9cEb ::text").getall()
        )
        address = response.css("div.fccl3c span::text").get("").strip()
        item["Address"] = address
        item["Industry"] = response.meta["Industry"]
        if address:
            address = address.replace("USA", "United States").replace(
                "US", "United States"
            )
            try:
                parsed_address, address_type = usaddress.tag(address)
                item["Street"] = (
                    parsed_address.get("AddressNumber", "")
                    + " "
                    + parsed_address.get("StreetName", ""),
                )
                item["City"] = "".join(list(parsed_address.get("PlaceName", [""])))
                item["State"] = "".join(list(parsed_address.get("StateName", [""])))
                item["Zipcode"] = "".join(list(parsed_address.get("ZipCode", [""])))
                item["Country"] = parsed_address.get("CountryName", "")
                if not item["State"]:
                    for i in self.us_states.keys():
                        if i in address.lower():
                            item["State"] = self.us_states[i]
                        if (
                            "united states" in address.lower()
                            or "usa" in address.lower()
                            or "us" in address.lower()
                        ):
                            item["Country"] = "United States"
                else:
                    address = (
                        address.replace(item["City"], "")
                        .replace(item["State"] + " ", "")
                        .replace(item["Zipcode"], "")
                        .replace(item["Country"], "")
                        .replace(",", "")
                    )
                    item["Street"] = address.strip()
            except Exception as e:
                print(f"Error is {e}")
                item["Street"] = ""
                item["City"] = ""
                item["State"] = ""
                item["Zipcode"] = ""
                item["Country"] = ""

        item["Description"] = response.css("h3.NwfE3d+div::attr(data-long-text)").get(
            ""
        )
        item["Review_Type"] = response.meta["type"]
        item["Google_Map_Link"] = (
            response.css('a[aria-label="Directions"]::attr(href)').get("").strip()
        )
        sttr = '//script[contains(text(),"hash: {}")]/text()'.format("'3'")
        script = response.xpath(sttr).get("{}")
        days = [
            "Tuesday",
            "Wednesday",
            "Thursday",
            "Friday",
            "Saturday",
            "Sunday",
            "Monday",
        ]
        opening_time = []
        business_hours = {}
        found = False
        for day in days:
            match = re.search(f'"{day}",(.*?)false', script)
            if match:
                found = True
                hours = match.group(1).split('[["')[-1].split('"')[0]
                opening_time.append(f"{day}: {hours}")
                business_hours[day] = hours
        if not found:
            sttr = '//script[contains(text(),"hash: {}")]/text()'.format("'4'")
            script = response.xpath(sttr).get("{}")
            days = [
                "Tuesday",
                "Wednesday",
                "Thursday",
                "Friday",
                "Saturday",
                "Sunday",
                "Monday",
            ]
            opening_time = []
            for day in days:
                match = re.search(f'"{day}",(.*?)false', script)
                if match:
                    hours = match.group(1).split('[["')[-1].split('"')[0]
                    opening_time.append(f"{day}: {hours}")
                    business_hours[day] = hours
        item["Opening Hours"] = " / ".join(opening_time)
        item["Business_Hours"] = business_hours
        feature_id = response.meta.get("feature_id")
        data = deepcopy(self.img_data).format(quote(feature_id), quote(feature_id))
        yield scrapy.Request(
            self.image_url,
            headers=self.img_header,
            body=data,
            method="POST",
            callback=self.image,
            meta={"item": item, 'dont_retry': True},
            errback=self.error,
            cb_kwargs={"item": item},
        )

    def image(self, response, item):
        item = response.meta.get("item")
        data = str(response.body)
        pattern = r"https://lh5\.googleusercontent\.com/p.*?(?=\\\\)"
        matches = re.findall(pattern, data)
        item["Images"] = ", ".join(matches)
        latitude, longitude = self.get_lat_lng(item["Address"])
        if latitude and longitude:
            item["Lat"] = latitude
            item["Lon"] = longitude
            item.pop("Google_Map_Link")
            yield item
        else:
            try:
                if item["Google_Map_Link"]:
                    yield scrapy.Request(
                        url=item["Google_Map_Link"],
                        headers=self.headers,
                        callback=self.lon_lat,
                        meta={"item": item, 'dont_retry': True},
                        errback=self.error,
                        cb_kwargs={"item": item},
                    )
                else:
                    item.pop("Google_Map_Link")
                    yield item
            except:
                item.pop("Google_Map_Link")
                yield item

    def lon_lat(self, response, item):
        item = response.meta.get("item")
        values = (
            response.xpath("//script/text()")
            .re_first("window.APP_INITIALIZATION_STATE=(.*),")
            .split("]", 1)[0]
            .split(",")[1:]
        )
        item["Lat"] = values[0]
        item["Lon"] = values[-1]
        item.pop("Google_Map_Link")
        yield item

    def get_lat_lng(self, address):
        try:
            geolocator = Nominatim(user_agent="my_geocoder", timeout=30)
            location = geolocator.geocode(address)
            if location:
                return location.latitude, location.longitude
            else:
                return None, None
        except Exception as e:
            print(f"Exception Occur for lat, Long: {e}")
            return None, None
